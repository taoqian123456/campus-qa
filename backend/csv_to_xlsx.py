"""一次性工具：把评测 CSV 转成 xlsx（WPS/Excel 双击即开，格式不乱，方便打分）。
用法：venv\\Scripts\\python.exe csv_to_xlsx.py
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parent

# 表头 -> 列宽（字符）
COLUMNS = [
    ("question", 38),
    ("answer", 70),
    ("reference", 42),
    ("相关性评分(1-5)", 14),
    ("忠实度评分(1-5)", 14),
]
HEADER_FILL = PatternFill("solid", fgColor="4F7CFF")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SCORE_FILL = PatternFill("solid", fgColor="FFF7E6")  # 评分列浅黄底，一眼找到


def convert(csv_path: Path):
    with open(csv_path, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))

    wb = Workbook()
    ws = wb.active
    ws.title = "评测打分"

    for ci, (name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, row in enumerate(rows[1:], 2):  # 跳过表头
        for ci in range(1, 6):
            val = row[ci - 1] if ci - 1 < len(row) else ""
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if ci in (4, 5):  # 评分列
                cell.fill = SCORE_FILL
                cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.freeze_panes = "A2"
    wb.save(csv_path.with_suffix(".xlsx"))
    return len(rows) - 1


if __name__ == "__main__":
    targets = sorted(BASE.glob("eval_results_*.csv"))
    for p in targets:
        n = convert(p)
        print(f"已转换：{p.name} -> {p.stem}.xlsx（{n} 题）")
    print(f"\n共 {len(targets)} 个文件，用 WPS/Excel 双击打开 .xlsx 打分即可。")
