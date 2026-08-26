"""一次性工具：为评测回答生成"建议分"（辅助人工复核，最终分仍由人确认）。

建议分规则（可解释、可写进论文"自动预评分辅助人工复核"）：
- 走兜底（未检索到）：相关性=1（没回答问题）、忠实度=5（没有编造，诚实的回答）
- 正常回答：相关性 = 回答与标准答案要点的语义相似度（bge-m3 余弦）映射 1-5；
  忠实度 = 带引用标注("依据：资料") 4-5 分，无标注但内容具体 3-4 分，空泛 2-3 分
输出：*_复核.xlsx，含建议分 + 理由 + 空白的最终分列（人工确认时填）
"""
import csv
import sys
from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from qa.embeddings import get_embedder  # noqa: E402

BASE = Path(__file__).resolve().parent
EMBEDDER = get_embedder()

FALLBACK_MARKERS = ("没有在知识库中检索到", "没有检索到", "未找到相关信息", "非常抱歉")


def cos(a: str, b: str) -> float:
    va = np.array(EMBEDDER.embed_query(a), dtype=np.float32)
    vb = np.array(EMBEDDER.embed_query(b), dtype=np.float32)
    va = va / max(np.linalg.norm(va), 1e-12)
    vb = vb / max(np.linalg.norm(vb), 1e-12)
    return float(np.dot(va, vb))


def suggest(answer: str, reference: str):
    """返回 (建议相关性, 建议忠实度, 理由)。"""
    if any(m in answer for m in FALLBACK_MARKERS):
        return 1, 5, "未命中走兜底：未回答问题但未编造（忠实度给满分）"
    sim = cos(answer, reference)
    # 相似度映射：0.40 -> 1 分，0.70 -> 5 分（线性）
    corr = round(1 + 4 * max(0.0, min(1.0, (sim - 0.40) / 0.30)), 1)
    if "依据：资料" in answer or "（依据" in answer:
        faith = 5.0 if sim >= 0.55 else 4.0
        faith_reason = "带引用标注"
    else:
        faith = 4.0 if sim >= 0.55 else 3.0
        faith_reason = "无引用标注"
    reason = f"与标准答案语义相似度 {sim:.2f}（{corr} 分）；{faith_reason}"
    return corr, faith, reason


def main():
    targets = sorted(BASE.glob("eval_results_*.csv"))
    for csv_path in targets:
        with open(csv_path, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))

        wb = Workbook()
        ws = wb.active
        ws.title = "复核打分"
        headers = [
            ("题目", 36), ("AI 回答", 60), ("标准答案要点", 36),
            ("建议相关性(1-5)", 13), ("建议忠实度(1-5)", 13), ("建议理由", 30),
            ("最终相关性(1-5)", 13), ("最终忠实度(1-5)", 13),
        ]
        for ci, (name, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=name)
            cell.fill = PatternFill("solid", fgColor="4F7CFF")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(ci)].width = width

        for ri, row in enumerate(rows[1:], 2):
            question = row[0] if len(row) > 0 else ""
            answer = row[1] if len(row) > 1 else ""
            reference = row[2] if len(row) > 2 else ""
            corr, faith, reason = suggest(answer, reference)
            values = [question, answer, reference, corr, faith, reason, "", ""]
            for ci, v in enumerate(values, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if ci in (4, 5, 7, 8):
                    cell.alignment = Alignment(horizontal="center", vertical="top")
                    cell.fill = PatternFill("solid", fgColor="FFF7E6")

        ws.freeze_panes = "A2"
        out = csv_path.with_name(csv_path.stem + "_复核.xlsx")
        wb.save(out)
        print(f"已生成：{out.name}（{len(rows) - 1} 题）")


if __name__ == "__main__":
    main()
