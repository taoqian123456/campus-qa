"""一次性工具：把复核表填满——最终分 = 用户填的值（非空）或建议分（空）。
同时输出每组平均分汇总。运行后复核表即为最终评分表。
"""
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent
targets = sorted(BASE.glob("eval_results_*_复核.xlsx"))

summary = []
for p in targets:
    wb = load_workbook(p)
    ws = wb.active
    n = 0
    filled = 0
    corr_sum = 0.0
    faith_sum = 0.0
    for row in ws.iter_rows(min_row=2, values_only=False):
        corr_sug = row[3].value  # 建议相关性
        faith_sug = row[4].value  # 建议忠实度
        corr_final = row[6].value  # 最终相关性
        faith_final = row[7].value  # 最终忠实度
        n += 1
        # 空 = 采纳建议分；非空 = 用户已改（保留）
        vc = corr_final if corr_final is not None and str(corr_final).strip() != "" else corr_sug
        vf = faith_final if faith_final is not None and str(faith_final).strip() != "" else faith_sug
        if corr_final is not None and str(corr_final).strip() != "":
            filled += 1
        row[6].value = float(vc)
        row[7].value = float(vf)
        corr_sum += float(vc)
        faith_sum += float(vf)
    wb.save(p)
    summary.append((p.name, n, filled, round(corr_sum / n, 2), round(faith_sum / n, 2)))

out = ["=== 最终评分汇总（最终分 = 你改的分，空位 = 采纳建议分）==="]
out.append(f"{'文件':<52}{'题数':<5}{'你改过':<6}{'平均相关性':<9}{'平均忠实度'}")
for name, n, filled, c, f in summary:
    out.append(f"{name:<52}{n:<5}{filled:<6}{c:<9}{f}")
open(BASE / "评分汇总.txt", "w", encoding="utf-8").write("\n".join(out))
print("done")
